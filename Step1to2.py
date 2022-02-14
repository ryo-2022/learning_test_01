import pandas as pd
import openpyxl

#STEP1 Start
#CSVデータの読み込み
url = "https://opendata.city.minato.tokyo.jp/dataset/74c06ebb-47dd-4fe1-8ba7-a5be60d2a448/resource/fbd63677-3f3c-4a85-9595-7ecaab739afd/download/minatokushisetsujoho_kuyakusyo.csv"
df_kuyakusyo = pd.read_csv(url, encoding="utf-8")

#CSVデータをExcelへ書き出し
df_kuyakusyo.to_excel("C:\\bbt\\チャレンジ課題.xlsx")

#STEP2 Start
#データ加工
wb = openpyxl.load_workbook("C:\\bbt\\チャレンジ課題.xlsx")
sheet = wb["Sheet1"]
sheet.title = "港区_区役所一覧"

#表の整形（不要な行列の削除と列名の変更）
sheet.delete_cols(1,2)
sheet.delete_rows(9)
sheet.cell(row=1, column=1).value = "施設名"

#検索キー生成と挿入
sheet.insert_cols(1)
sheet.cell(row=1, column=1).value = "検索キー"
#最終行を取得
lastrow=sheet.max_row
print("lastrow" + str(lastrow))
for i in range(lastrow - 1) :
    val1 = sheet.cell(row=i+2, column=2).value
    val2 = sheet.cell(row=i+2, column=3).value
    sheet.cell(row=i+2, column=1).value = val1 + str(val2)

#Excel の保存   
wb.save("C:\\bbt\\チャレンジ課題.xlsx")
wb.close()
